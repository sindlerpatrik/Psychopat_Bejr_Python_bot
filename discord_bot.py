from discord.ext import commands
import discord, requests
from datetime import datetime
from openpyxl import load_workbook

BOT_TOKEN = "YOUR_TOKEN"
CHANNEL_ID = your channel ID

bot = commands.Bot(command_prefix = "/", intents = discord.Intents.all())
    
@bot.event
async def on_ready():
    print("Nazdar ty močko, ty chcanko zkurvená!!")
    channel = bot.get_channel(CHANNEL_ID)        
    
    await bot.tree.sync()        
    await channel.send("HEHEHE, KOPULEJŠN")
    
@bot.tree.command(name = "weather", description = "psychopat ti záhlásí počasí")
async def w(interaction : discord.Interaction, city : str):
    try:
        Response = requests.get(f"https://api.openweathermap.org/data/2.5/forecast?q={city}&appid=88338ec27e1c8ca89107de0bc0c4e561&cnt=3&units=metric")
        Response.raise_for_status()
        
        actors = ""
        
        data = Response.json()
        
        date = datetime.now()
        formatted_time = date.strftime("%d-%m-%Y %H:%M:%S")
        
        weather_data_list = [
        formatted_time,
        data["city"]["name"],
        data["city"]["country"],
        data["list"][0]["weather"][0]["main"],
        data["list"][0]["weather"][0]["description"],
        data["list"][0]["main"]["temp"],
        data["list"][0]["main"]["humidity"],
        round(data["list"][0]["wind"]["speed"] * 3.6, 2),
        ]
        
        timezone = data["city"]["timezone"]
        city_name = weather_data_list[1]
        country = weather_data_list[2]
        weather = weather_data_list[3]
        weather_description = weather_data_list[4]
        temperature = weather_data_list[5]
        humidity = weather_data_list[6]
        wind_speed = weather_data_list[7]
        
        time_difference = round((timezone / 3600))
        
        if time_difference >= 0:
            time_difference = "+" + str(time_difference)
        
        wb = load_workbook("weather_data.xlsx")

        ws = wb.active

        row = ws.max_row + 1
        
        if ws["A2"].value == "":
            row = 1

        cells = ["A", "B", "C", "D", "E", "F", "G", "H"]

        for index, cell in enumerate(cells):
            ws[f"{cell}{row}"] = weather_data_list[index]

        wb.save("weather_data.xlsx")
        
        if temperature <= 0:
            actors = "Tyvole, tam je zima jako kráva. Dej Bůh sílu všem hercům tam venku.🙏🥶"
            
        elif 0 < temperature < 20:
            actors = "Není to pro herce sice nejlepší, ale aspoň nezmrznou ty močky.🙂"
        
        else:
            actors = "Tyvole tam je teplíčko meheheh. To se každý herec musí mít nádherně.🥵"

        formatted_info = f"{formatted_time} ({time_difference}h UCT)\n**{city_name}, {country}**\n\nTady máš jaký je počáskos ty močále\n\n{actors}\n\n**Počáskos**: {weather}\n**Přesnější popiskos**: {weather_description}\n**Teplotkos**: {temperature}°C\n**Vlhostos**: {humidity}%\n**Rychlostos větros**: {wind_speed}km/h"
        await interaction.response.send_message(formatted_info)

    except requests.exceptions.HTTPError as http_err:
        data = Response.json()
        error_code = data["cod"]
        message = data["message"]
        await interaction.response.send_message(f"Error kodikos: {error_code}\nDůvodos: {message}\nChybičkos při komunikaci ty močko, buď jsi zadal píčovinu nebo špatně název města nebo obce!! Zkus to třeba s diakritikou nebo in ingliš ty srágoro!!")

    except Exception as e:
        await interaction.response.send_message(f"Nočekávaná chybkos, něco se totálně posralo: {e}")

@bot.tree.command(name = "ping", description = "ukáže se ti botův pingos")
async def ping(interaction : discord.Interaction):
    bot_latency = round(bot.latency*1000)
    await interaction.response.send_message(f"Pong!Zmrde! {bot_latency}ms")

@bot.tree.command(name = "avatar", description = "avatar herce")
async def avatar(interaction : discord.Interaction, member : discord.Member):
    await interaction.response.send_message(member.display_avatar)

if __name__ == '__main__':
    bot.run(BOT_TOKEN)